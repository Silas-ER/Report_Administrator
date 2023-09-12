#############################################################################################################
####################################### FUNÇÃO PARA CRIAÇÃO DO CABEÇALHO ####################################
#############################################################################################################

from openpyxl.styles import Font, Alignment

def criar_cabecalho(sheet, linha, data):

    sheet.merge_cells("A{}:J{}".format(linha, linha))
    sheet["A{}".format(linha)].value = "Anexo II"
    sheet["A{}".format(linha)].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A{}:J{}".format(linha + 1, linha + 1))
    sheet["A{}".format(linha + 1)].value = "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO"
    sheet["A{}".format(linha + 1)].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A{}:J{}".format(linha + 2, linha + 2))
    sheet["A{}".format(linha + 2)].value = "SECRETARIA DE AQUICULTURA E PESCA"
    sheet["A{}".format(linha + 2)].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A{}:C{}".format(linha + 3, linha + 3))
    sheet["A{}".format(linha + 3)].value = "NOME DA EMPRESA PESQUEIRA:"

    sheet["D{}".format(linha + 3)].value = "MAR EXPORTADORA DE PRODUTOS DO MAR LTDA"
    empresa = sheet.cell(row=4, column=4)
    empresa.font = Font(underline="single")
    sheet.merge_cells("D{}:J{}".format(linha + 3, linha + 3))

    sheet.merge_cells("A{}:C{}".format(linha + 4, linha + 4))
    sheet["A{}".format(linha + 4)].value = "ENDEREÇO E MUNICÍPIO:"

    sheet["D{}".format(linha + 4)].value = "RUA TAL, 128, TAL , NATAL-RN"
    endereco = sheet.cell(row=5, column=4)
    endereco.font = Font(underline="single")
    sheet.merge_cells("D{}:H{}".format(linha + 4, linha + 4))

    sheet["I{}".format(linha + 4)].value = "CEP:"
    sheet["J{}".format(linha + 4)].value = "00.000-000"
    cep = sheet.cell(row=5, column=10)
    cep.font = Font(underline="single")

    sheet["A{}".format(linha + 5)].value = "RGP EMPRESA PESQUEIRA:"
    sheet.merge_cells("A{}:C{}".format(linha + 5, linha + 5))
    sheet.merge_cells("D{}:E{}".format(linha + 5, linha + 5))

    sheet["F{}".format(linha + 5)].value = "CNPJ:"
    sheet["G{}".format(linha + 5)].value = "00.000.000/0000-00"
    cnpj = sheet.cell(row=6, column=7)
    cnpj.font = Font(underline="single")
    sheet.merge_cells("G{}:J{}".format(linha + 5, linha + 5))

    sheet["A{}".format(linha + 6)].value = "NOME DO RESPONSÁVEL LEGAL:"
    sheet.merge_cells("A{}:C{}".format(linha + 6, linha + 6))    
    sheet.merge_cells("D{}:J{}".format(linha + 6, linha + 6))
    sheet["A{}".format(linha + 7)].value = "CPF RESPONSÁVEL LEGAL:"
    sheet.merge_cells("A{}:C{}".format(linha + 7, linha + 7)) 
    sheet.merge_cells("D{}:J{}".format(linha + 7, linha + 7))
    sheet["A{}".format(linha + 8)].value = "DATA: {}".format(data)
    sheet.merge_cells("A{}:C{}".format(linha + 8, linha + 8)) 
    sheet.merge_cells("D{}:J{}".format(linha + 8, linha + 8))

    #filtros e cabeçalhos de coluna
    sheet["A{}".format(linha + 9)].value = "ENTRADA"
    sheet.merge_cells("A{}:B{}".format(linha + 9, linha + 9))
    sheet["A{}".format(linha + 9)].alignment = Alignment(horizontal="center")

    sheet["C{}".format(linha + 9)].value = "ESPÉCIE DE LAGOSTA E QUANTIDADE"
    sheet.merge_cells("C{}:H{}".format(linha + 9, linha + 9))
    sheet["C{}".format(linha + 9)].alignment = Alignment(horizontal="center")

    sheet["I{}".format(linha + 9)].value = "NF"
    sheet["I{}".format(linha + 9)].alignment = Alignment(horizontal="center")

    sheet["J{}".format(linha + 9)].value = "OBSERVAÇÃO"
    sheet["J{}".format(linha + 9)].alignment = Alignment(horizontal="center")