##################################################################################################
#################### FUNÇÃO PARA ATRIBUIR AS FÓRMULAS AS COLUNAS SELECIONADAS ####################
##################################################################################################

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def escrever_formulas(sheet, linha_init):
    for col, divisor in [(4, 0.6), (6, 0.45), (8, 0.45)]:
        col_atual = get_column_letter(col)
        col_anterior = get_column_letter(col - 1)
        
        formula = f'={col_anterior}{linha_init}/{divisor}'
        
        cell = sheet[f'{col_atual}{linha_init}']
        cell.value = formula
        cell.number_format = '0.00'